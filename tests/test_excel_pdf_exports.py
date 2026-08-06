"""Exportaciones documentales: contenido real, seguridad y relectura."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from horizun_pbi_mcp.pbip import project_locator
from horizun_pbi_mcp.powerbi import desktop_capture
from horizun_pbi_mcp.powerbi.errors import ValidationError
from horizun_pbi_mcp.services import exporting


@pytest.fixture
def proyecto(session, sample_pbip):
    project_locator.open_project(session, str(sample_pbip))
    return session


def test_excel_exporta_modelo_reporte_y_auditoria(proyecto):
    result = exporting.export_excel(
        proyecto, source="pbip", include_report=True, include_audit=True)

    path = Path(result["output_path"])
    assert path.is_file() and path.suffix == ".xlsx"
    assert result["verified"] is True
    assert {"Resumen", "Tablas", "Columnas", "Medidas", "Relaciones",
            "Paginas", "Auditoria"} <= set(result["sheets"])

    workbook = load_workbook(path, read_only=True, data_only=False)
    assert workbook["Resumen"]["B6"].value == 1  # tablas
    assert workbook["Tablas"]["A2"].value == "Ventas"
    assert workbook["Medidas"].max_row >= 2
    workbook.close()


def test_excel_no_interpreta_nombres_como_formulas():
    assert exporting._json_cell("=HYPERLINK(\"https://example.invalid\")").startswith("'")
    assert exporting._json_cell("@SUM(A1:A2)").startswith("'")
    assert exporting._json_cell(-5) == -5, "un numero negativo sigue siendo numero"


def test_excel_hace_unicos_los_encabezados_dax():
    assert exporting._unique_headers(["Importe", "importe", "", "Importe"]) == [
        "Importe", "importe_2", "Columna_3", "Importe_3"]


def test_excel_rechaza_filas_dax_que_no_corresponden_a_encabezados(
        proyecto, monkeypatch):
    context = exporting.ExportContext(
        source="live", model={"tables": [], "measures": [], "relationships": []},
        query_result={"columns": ["A"], "rows": [[1, 2]], "row_count": 1,
                      "truncated": False})
    monkeypatch.setattr(exporting, "_collect_context", lambda *_a, **_k: context)

    with pytest.raises(ValidationError, match="distinta cantidad"):
        exporting.export_excel(proyecto, source="live")


def test_pdf_escapa_markup_de_nombres():
    assert exporting._pdf_text("A&B <Total>") == "A&amp;B &lt;Total&gt;"


def test_excel_no_sobrescribe_un_nombre_existente(proyecto):
    exporting.export_excel(
        proyecto, source="pbip", include_audit=False, file_name="estable.xlsx")
    with pytest.raises(ValidationError):
        exporting.export_excel(
            proyecto, source="pbip", include_audit=False,
            file_name="estable.xlsx")


def test_query_dax_exige_fuente_live(proyecto):
    with pytest.raises(ValidationError):
        exporting.export_excel(
            proyecto, source="pbip", query='EVALUATE ROW("x", 1)',
            include_audit=False)


def test_pdf_incluye_captura_y_se_reabre(proyecto, tmp_path):
    # PNG sintetico de cuatro colores; no depende de Pillow ni de Desktop.
    bgra = (b"\x00\x00\xff\xff" + b"\x00\xff\x00\xff" +
            b"\xff\x00\x00\xff" + b"\xff\xff\xff\xff")
    capture = tmp_path / "tablero.png"
    capture.write_bytes(desktop_capture._encode_png(2, 2, bgra))

    result = exporting.generate_pdf_report(
        proyecto, source="pbip", report_type="technical",
        title="Reporte de prueba", capture_paths=[str(capture)],
        include_audit=True)

    path = Path(result["output_path"])
    reader = PdfReader(str(path))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert result["verified"]["logical"] is True
    assert result["captures_included"] == 1
    assert len(reader.pages) == result["pages"] >= 2
    assert "Reporte de prueba" in text
    assert "Capturas de los tableros" in text


def test_etiqueta_de_objeto_nombra_cada_tipo_de_hallazgo():
    paginas = {"pg1": "Matriz de Riesgos"}
    etiqueta = exporting._finding_object_label

    assert etiqueta({"kind": "measure", "name": "Total", "table": "Ventas"}) == \
        "Medida Total (Ventas)"
    assert etiqueta({"kind": "column", "name": "Ventas[Monto]"}) == \
        "Columna Ventas[Monto]"
    assert etiqueta({"kind": "relationship", "from": "A", "to": "B"}) == \
        "Relacion A -> B"
    assert etiqueta({"kind": "visual", "id": "abc123", "type": "textbox",
                     "page": "pg1"}, paginas) == \
        "Visual textbox abc123 - pagina Matriz de Riesgos"
    assert etiqueta({"kind": "visual_group", "visuals": ["a1", "b2"],
                     "page": "pg1"}, paginas) == \
        "Visuales a1, b2 - pagina Matriz de Riesgos"
    # Un grupo con toda la pagina dentro no puede reventar la fila del PDF.
    assert etiqueta({"kind": "visual_group",
                     "visuals": [f"v{n}" for n in range(7)]}) == \
        "Visuales v0, v1, v2, v3, (+3 mas)"
    assert etiqueta({"kind": "page", "page": "pg1", "name": "P1"}) == "Pagina P1"
    assert etiqueta({"kind": "page", "page": "pg1"}, paginas) == \
        "Pagina Matriz de Riesgos"
    assert etiqueta({"kind": "model"}) == "Modelo"
    assert etiqueta(None) == "" and etiqueta({}) == ""


def test_texto_del_hallazgo_no_vuelca_json_crudo():
    # Los motores de auditoria no producen `message`: solo `evidence`.
    assert exporting._finding_message(
        {"evidence": {"visual_count": 13, "threshold": 12}}) == \
        "visual_count: 13; threshold: 12"
    assert exporting._finding_message(
        {"evidence": {"visuals": ["a", "b"]}}) == "visuals: a, b"
    assert exporting._finding_message({"evidence": {}}) == ""
    # Ni `None` ni `False` de Python deben llegar a un reporte que lee alguien.
    assert exporting._finding_message(
        {"evidence": {"title": None, "has_description": False}}) == \
        "title: (vacio); has_description: no"
    assert exporting._finding_message(
        {"message": "Explicito", "evidence": {"x": 1}}) == "Explicito"


def test_pdf_de_auditoria_nombra_el_objeto_de_cada_hallazgo(proyecto):
    result = exporting.generate_pdf_report(
        proyecto, source="pbip", report_type="audit", include_audit=True)

    reader = PdfReader(result["output_path"])
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert result["findings_included"] >= 1
    # El reporte de auditoria no lista las tablas del modelo: si aparece
    # "Ventas" es porque el hallazgo de la medida por fin dice de quien habla.
    assert "Ventas" in text and "Total" in text
    assert '{"' not in text, "la evidencia no puede llegar como JSON al PDF"


def test_pdf_rechaza_capturas_ajenas_a_imagen(proyecto, tmp_path):
    bad = tmp_path / "captura.txt"
    bad.write_text("no es una imagen", encoding="utf-8")
    with pytest.raises(ValidationError):
        exporting.generate_pdf_report(
            proyecto, source="pbip", capture_paths=[str(bad)])


def test_pdf_rechaza_archivo_png_corrupto(proyecto, tmp_path):
    bad = tmp_path / "captura.png"
    bad.write_bytes(b"esto no es png")
    with pytest.raises(ValidationError, match="no es una imagen valida"):
        exporting.generate_pdf_report(
            proyecto, source="pbip", capture_paths=[str(bad)])


def test_publicacion_exclusiva_limpia_la_reserva_si_falla(monkeypatch, tmp_path):
    target = tmp_path / "fallido.xlsx"
    monkeypatch.setattr(
        exporting, "durable_write",
        lambda *_a, **_k: (_ for _ in ()).throw(OSError("corte inyectado")))

    with pytest.raises(OSError, match="corte inyectado"):
        exporting._publish_new_file(target, b"contenido")

    assert not target.exists()


def test_nombres_de_salida_no_admiten_rutas(proyecto, tmp_path):
    with pytest.raises(Exception):
        exporting.export_excel(
            proyecto, source="pbip", include_audit=False,
            file_name="../fuera.xlsx")
    assert not (tmp_path / "fuera.xlsx").exists()
