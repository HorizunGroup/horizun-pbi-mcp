"""Export del CONTENIDO: reconstruccion de la consulta y verdad del dato."""
from __future__ import annotations

import json
from pathlib import Path

import pytest

from horizun_pbi_mcp.pbip import filter_builder, filter_reader, project_locator
from horizun_pbi_mcp.powerbi import dax_runner
from horizun_pbi_mcp.powerbi.errors import ValidationError
from horizun_pbi_mcp.services import content_export, content_query

# --- relaciones de juguete: Medidas (muchos) -> Riesgos (uno) ---------------
RELACIONES = [{"name": "r1", "from_table": "Medidas", "from_column": "ID",
               "to_table": "Riesgos", "to_column": "ID", "is_active": True}]


def _proyeccion_columna(entidad, propiedad, **extra):
    return {"kind": "column", "entity": entidad, "property": propiedad,
            "ref": f"{entidad}[{propiedad}]", **extra}


def _proyeccion_medida(entidad, nombre, **extra):
    return {"kind": "measure", "entity": entidad, "property": nombre,
            "ref": f"{entidad}[{nombre}]", **extra}


# ===========================================================================
# Lectura de filtros
# ===========================================================================

def test_filtro_categorico_se_traduce_con_sus_valores():
    entrada = filter_builder.build_filter(
        {"field": "Riesgos[Estado]", "values": ["Abierto", "En curso"]})
    leido = filter_reader.leer_filtro(entrada, scope="pagina")

    assert leido["state"] == "applied"
    assert leido["field"] == "Riesgos[Estado]"
    assert leido["values"] == ["Abierto", "En curso"]
    assert leido["exclude"] is False


def test_filtro_de_exclusion_conserva_el_signo():
    entrada = filter_builder.build_filter(
        {"field": "Riesgos[Estado]", "values": ["Cerrado"], "exclude": True})
    leido = filter_reader.leer_filtro(entrada, scope="visual")

    assert leido["state"] == "applied" and leido["exclude"] is True


def test_campo_en_el_panel_sin_seleccion_no_es_un_filtro():
    entrada = filter_builder.build_filter({"field": "Riesgos[Estado]"})
    assert filter_reader.leer_filtro(entrada, scope="pagina")["state"] == "unset"


def test_filtro_sobre_medida_no_se_traduce():
    entrada = {"field": {"Measure": {"Expression": {"SourceRef": {"Entity": "M"}},
                                     "Property": "Total"}},
               "type": "Advanced", "filter": {"Where": [{"Condition": {}}]}}
    leido = filter_reader.leer_filtro(entrada, scope="visual")

    assert leido["state"] == "untranslated"
    assert "medida" in leido["reason"]


def test_in_sobre_dos_columnas_no_equivale_a_una_lista():
    campo = {"Column": {"Expression": {"SourceRef": {"Source": "r"}},
                        "Property": "Estado"}}
    entrada = {"field": {"Column": {"Expression": {"SourceRef": {"Entity": "Riesgos"}},
                                    "Property": "Estado"}},
               "type": "Categorical",
               "filter": {"Where": [{"Condition": {"In": {
                   "Expressions": [campo, campo],
                   "Values": [[{"Literal": {"Value": "'a'"}},
                               {"Literal": {"Value": "'b'"}}]]}}}]}}
    assert filter_reader.leer_filtro(entrada, scope="visual")["state"] == "untranslated"


def test_literales_vuelven_a_su_tipo():
    assert filter_reader._parse_literal("'Ana''s'") == "Ana's"
    assert filter_reader._parse_literal("5L") == 5
    assert filter_reader._parse_literal("true") is True
    assert filter_reader._parse_literal("null") is None


def test_resumen_marca_como_no_fiable_lo_que_no_se_tradujo():
    filtros = [{"state": "applied"}, {"state": "untranslated"}]
    assert filter_reader.resumen(filtros)["trustworthy"] is False
    assert filter_reader.resumen([{"state": "applied"}])["trustworthy"] is True


# ===========================================================================
# Sintesis de la consulta
# ===========================================================================

def test_referencias_escapan_comillas_y_corchetes():
    assert content_query.columna_dax("O'Brien[Col]") == "'O''Brien'[Col]"
    assert content_query.medida_dax("Ventas[Total]") == "[Total]"


def test_los_filtros_van_dentro_de_summarizecolumns():
    dax = content_query.construir_dax(
        ["Riesgos[Estado]"], [{"alias": "Total", "expr": "[Total]"}],
        [{"field": "Riesgos[Fase]", "values": ["Obra"]}])

    assert dax.startswith("EVALUATE")
    assert "FILTER(ALL('Riesgos'[Fase])" in dax
    # DAX prohibe SUMMARIZECOLUMNS dentro de un contexto modificado por
    # CALCULATE: envolverlo en CALCULATETABLE falla en el motor.
    assert "CALCULATETABLE" not in dax


def test_la_tabla_de_hechos_es_la_del_lado_muchos():
    assert content_query.tabla_de_hechos(["Riesgos", "Medidas"], RELACIONES) == "Medidas"
    assert content_query.tabla_de_hechos(["Medidas"], RELACIONES) == "Medidas"


def test_sin_tabla_de_hechos_no_se_exporta_un_producto_cartesiano():
    visual = {"id": "v1", "type": "tableEx", "title": "Cruce",
              "fields": {"Rows": [_proyeccion_columna("A", "X"),
                                  _proyeccion_columna("B", "Y")]}}
    plan = content_query.plan_de_visual(visual, (), [])

    assert plan["exportable"] is False
    assert "producto cartesiano" in plan["reason"]


def test_columnas_de_dos_tablas_agregan_la_medida_de_existencia():
    visual = {"id": "v1", "type": "tableEx", "title": "Plan",
              "fields": {"Rows": [_proyeccion_columna("Riesgos", "ID"),
                                  _proyeccion_columna("Medidas", "Accion")]}}
    plan = content_query.plan_de_visual(visual, (), RELACIONES)

    assert plan["exportable"] is True
    aux = [m for m in plan["measures"] if m.get("aux")]
    assert len(aux) == 1
    # Comprobado contra el motor: sin esto son 400 filas donde el visual
    # muestra 20.
    assert aux[0]["expr"] == "CALCULATE(COUNTROWS('Medidas'))"
    assert content_query.ALIAS_EXISTENCIA in plan["dax"]


def test_una_sola_tabla_no_necesita_andamiaje():
    visual = {"id": "v1", "type": "barChart", "title": "Por fase",
              "fields": {"Category": [_proyeccion_columna("Riesgos", "Fase")],
                         "Y": [_proyeccion_medida("Riesgos", "Total")]}}
    plan = content_query.plan_de_visual(visual, (), RELACIONES)

    assert plan["exportable"] is True
    assert not any(m.get("aux") for m in plan["measures"])


def test_agregacion_implicita_se_traduce_por_su_nombre():
    visual = {"id": "v1", "type": "card", "title": "Monto",
              "fields": {"Values": [_proyeccion_columna(
                  "Ventas", "Monto", aggregation=0, queryRef="Sum(Ventas.Monto)")]}}
    plan = content_query.plan_de_visual(visual, (), RELACIONES)

    assert plan["exportable"] is True
    assert plan["measures"][0]["expr"] == "SUM('Ventas'[Monto])"


def test_agregacion_desconocida_declina_en_vez_de_adivinar():
    visual = {"id": "v1", "type": "card", "title": "Raro",
              "fields": {"Values": [_proyeccion_columna(
                  "Ventas", "Monto", aggregation=42)]}}
    plan = content_query.plan_de_visual(visual, (), RELACIONES)

    assert plan["exportable"] is False and "Agregacion" in plan["reason"]


def test_nombre_y_codigo_de_agregacion_que_no_coinciden_declinan():
    visual = {"id": "v1", "type": "card", "title": "Contradictorio",
              "fields": {"Values": [_proyeccion_columna(
                  "Ventas", "Monto", aggregation=0, queryRef="Max(Ventas.Monto)")]}}
    plan = content_query.plan_de_visual(visual, (), RELACIONES)

    assert plan["exportable"] is False
    assert "queryRef" in plan["reason"]


def test_un_cuadro_de_texto_no_es_un_dato_vacio():
    plan = content_query.plan_de_visual(
        {"id": "t1", "type": "textbox", "title": "", "fields": {}}, (), ())
    assert plan["exportable"] is False and "no consulta datos" in plan["reason"]


def test_top_n_no_ordena_por_la_medida_de_existencia():
    with pytest.raises(ValidationError, match="top_n"):
        content_query.plan_declarado(
            {"name": "Top", "rows": ["Riesgos[ID]", "Medidas[Accion]"],
             "top_n": 5}, RELACIONES)


def test_consulta_declarada_por_el_cliente():
    plan = content_query.plan_declarado(
        {"name": "Costo por fase", "rows": ["Riesgos[Fase]"],
         "values": ["Costo"], "filters": [{"field": "Riesgos[Estado]",
                                           "values": ["Abierto"]}],
         "top_n": 3}, RELACIONES)

    assert plan["title"] == "Costo por fase"
    assert "TOPN(3" in plan["dax"] and "[Costo]" in plan["dax"]


# ===========================================================================
# Export
# ===========================================================================

@pytest.fixture
def proyecto(session, sample_pbip):
    """El .pbip de siempre, con dos visuales y un cuadro de texto."""
    project_locator.open_project(session, str(sample_pbip))
    paginas = sample_pbip.parent / "MyReport.Report" / "definition" / "pages"
    visuales = paginas / "pg1" / "visuals"

    def escribir(vid: str, cuerpo: dict) -> None:
        (visuales / vid).mkdir(parents=True, exist_ok=True)
        (visuales / vid / "visual.json").write_text(
            json.dumps(cuerpo), encoding="utf-8")

    escribir("v_barras", {
        "name": "v_barras",
        "position": {"x": 0, "y": 0, "width": 300, "height": 200, "z": 0},
        "visual": {"visualType": "barChart", "query": {"queryState": {
            "Category": {"projections": [{"field": {"Column": {
                "Expression": {"SourceRef": {"Entity": "Ventas"}},
                "Property": "Monto"}}, "nativeQueryRef": "Monto"}]},
            "Y": {"projections": [{"field": {"Measure": {
                "Expression": {"SourceRef": {"Entity": "Ventas"}},
                "Property": "Total"}}, "nativeQueryRef": "Total"}]}}}},
        "filterConfig": {"filters": [filter_builder.build_filter(
            {"field": "Ventas[Monto]", "values": [10]})]},
    })
    escribir("v_texto", {
        "name": "v_texto",
        "position": {"x": 0, "y": 210, "width": 300, "height": 60, "z": 1},
        "visual": {"visualType": "textbox"},
    })
    return session


def test_la_seleccion_separa_lo_exportable_de_lo_que_no(proyecto):
    resuelto = content_export.resolver_seleccion(proyecto, {"pages": ["P1"]})

    assert len(resuelto["plans"]) == 1
    assert resuelto["plans"][0]["visual_type"] == "barChart"
    assert len(resuelto["skipped"]) == 1
    assert resuelto["skipped"][0]["visual_type"] == "textbox"


def test_la_pagina_se_puede_pedir_por_su_nombre_visible_o_su_id(proyecto):
    por_nombre = content_export.resolver_seleccion(proyecto, {"pages": ["P1"]})
    por_id = content_export.resolver_seleccion(proyecto, {"pages": ["pg1"]})
    assert [p["dax"] for p in por_nombre["plans"]] == [p["dax"] for p in por_id["plans"]]


def test_una_pagina_que_no_existe_dice_cuales_hay(proyecto):
    with pytest.raises(ValidationError, match="No existe la pagina"):
        content_export.resolver_seleccion(proyecto, {"pages": ["Fantasma"]})


def test_dry_run_devuelve_el_dax_sin_tocar_el_motor(proyecto, monkeypatch):
    def prohibido(*_a, **_k):                  # el motor no se toca
        raise AssertionError("dry_run no puede consultar el modelo")

    monkeypatch.setattr(dax_runner, "run_dax", prohibido)
    salida = content_export.export_content(
        proyecto, select={"pages": ["P1"]}, dry_run=True)

    assert salida["dry_run"] is True and salida["outputs"] == []
    assert salida["queries"][0]["dax"].startswith("EVALUATE")
    assert salida["queries"][0]["filters_applied"][0]["field"] == "Ventas[Monto]"


def test_el_filtro_del_visual_viaja_a_la_consulta(proyecto):
    salida = content_export.export_content(
        proyecto, select={"pages": ["P1"]}, dry_run=True)
    assert "FILTER(ALL('Ventas'[Monto])" in salida["queries"][0]["dax"]


def test_un_modelo_sin_datos_no_se_exporta_en_blanco(proyecto, monkeypatch, tmp_path):
    """Un .pbip recien abierto responde a las consultas y devuelve cero filas."""
    monkeypatch.setattr(content_export, "estado_de_datos",
                        lambda _s: {"partitions": 2, "ready": 0, "processed": False})
    monkeypatch.setattr(content_export, "_modelo_en_vivo", lambda *a, **k: [])
    monkeypatch.setattr(dax_runner, "run_dax", lambda *_a, **_k: {
        "columns": [], "rows": [], "row_count": 0, "truncated": False})

    with pytest.raises(ValidationError, match="SIN DATOS"):
        content_export.export_content(proyecto, select={"pages": ["P1"]})

    assert not list((tmp_path).rglob("*.xlsx")), "no se publica nada"


def test_las_fechas_no_llegan_en_formato_de_maquina():
    # Asi las devuelve el motor: round-trip de .NET.
    assert content_export.valor_legible("2026-09-11T00:00:00.0000000") == "2026-09-11"
    assert content_export.valor_legible("2026-09-11T14:30:00") == "2026-09-11 14:30:00"
    assert content_export.valor_legible("R-01") == "R-01"
    assert content_export.valor_legible(35) == 35


def test_los_encabezados_pierden_la_tabla_salvo_cuando_hay_ambiguedad():
    assert content_export.encabezados_legibles(
        ["Riesgos[Fase]", "[Total Riesgos]"]) == ["Fase", "Total Riesgos"]
    # Dos columnas 'Nombre' de tablas distintas conservan la forma larga.
    assert content_export.encabezados_legibles(
        ["A[Nombre]", "B[Nombre]", "A[Fecha]"]) == ["A[Nombre]", "B[Nombre]", "Fecha"]


def test_el_andamiaje_no_llega_al_archivo():
    plan = {"measures": [{"alias": "Total"}, {"alias": content_query.ALIAS_EXISTENCIA,
                                              "aux": True}]}
    columnas = ["Riesgos[ID]", "[Total]", f"[{content_query.ALIAS_EXISTENCIA}]"]
    filas = [["R1", 5, 1], ["R2", 7, 1]]

    limpias, datos = content_export._sin_auxiliares(plan, columnas, filas)

    assert limpias == ["Riesgos[ID]", "[Total]"]
    assert datos == [["R1", 5], ["R2", 7]]


def test_el_excel_lleva_los_datos_y_declara_sus_filtros(proyecto, monkeypatch):
    monkeypatch.setattr(content_export, "_asegurar_modelo", lambda *a, **k: [])
    monkeypatch.setattr(dax_runner, "run_dax", lambda *_a, **_k: {
        "columns": ["Ventas[Monto]", "[Total]"],
        "rows": [[10, 100], [20, 200]], "row_count": 2, "truncated": False,
        "elapsed_ms": 1.0})

    salida = content_export.export_content(
        proyecto, select={"pages": ["P1"]}, format="xlsx")

    from openpyxl import load_workbook
    ruta = Path(salida["outputs"][0]["output_path"])
    libro = load_workbook(ruta)
    assert "Contenido" in libro.sheetnames
    hoja = libro[[n for n in libro.sheetnames if n != "Contenido"][0]]
    celdas = [[c for c in fila] for fila in hoja.iter_rows(values_only=True)]
    plano = [str(c) for fila in celdas for c in fila if c is not None]

    valores = [c for fila in celdas for c in fila if c is not None]
    assert salida["queries"][0]["rows"] == 2
    assert any("Ventas[Monto] esta en (10)" in v for v in plano), \
        "cada hoja declara con que filtros se saco"
    assert ["Monto", "Total"] == [
        c for c in celdas[len(celdas) - 3] if c is not None], "encabezados del dato"
    assert 100 in valores and 200 in valores, "las filas del motor llegan al archivo"
