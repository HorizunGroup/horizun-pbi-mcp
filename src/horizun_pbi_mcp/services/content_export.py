"""Exportar el CONTENIDO de un informe: los datos, no los metadatos.

`exporting` documenta el proyecto -que tablas hay, que visuales, que
hallazgos-. Esto exporta lo que el tablero MUESTRA: la tabla que hay detras
de cada visual, o la que el cliente declare.

Tres cosas que este modulo se niega a hacer, porque son las tres formas de
entregar un archivo que miente:

1. **Exportar sin datos.** Un .pbip recien abierto en Desktop trae el modelo
   SIN procesar: las consultas responden vacio y el Excel sale en blanco sin
   un solo error. Se comprueba el estado de las particiones antes de
   consultar y se rechaza con instrucciones en vez de publicar el vacio.
2. **Ignorar un filtro que no se supo traducir.** Cada hoja declara sus
   filtros aplicados y, sobre todo, los que NO se pudieron aplicar.
3. **Inventarle una consulta a un visual que no la tiene.** Los que no son
   tabulares se listan aparte, con el motivo, en vez de salir vacios.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from horizun_pbi_mcp.config import Session
from horizun_pbi_mcp.logging_config import get_logger
from horizun_pbi_mcp.pbip import filter_reader, pbir_reader
from horizun_pbi_mcp.powerbi import dax_runner
from horizun_pbi_mcp.powerbi.errors import ValidationError
from horizun_pbi_mcp.services import content_query, exporting
from horizun_pbi_mcp.utils.json_utils import read_json

log = get_logger("content_export")

#: Tope de consultas por export. No es una cifra magica: cada una es un viaje
#: al motor, y un informe grande con `pages` sin acotar puede lanzar cientos
#: sin que nadie lo haya pedido explicitamente.
MAX_CONSULTAS = 60

#: Hueco de una celda. Un guion suelto lo neutraliza el guardia de
#: formulas de Excel y llega al cliente como `'-`.
SIN_DATO = "n/a"


def _pagina_por_nombre(active, referencia: str) -> Dict[str, Any]:
    """Acepta el id interno de la pagina o su nombre visible."""
    paginas = pbir_reader.list_pages(active)
    texto = str(referencia).strip()
    for p in paginas:
        if p.get("name") == texto or (p.get("display_name") or "") == texto:
            return p
    disponibles = [p.get("display_name") or p.get("name") for p in paginas]
    raise ValidationError(
        f"No existe la pagina '{referencia}' en el informe activo.",
        details={"requested": referencia, "available": disponibles})


def _filtros_de_pagina(active, page_id: str) -> List[Dict[str, Any]]:
    ruta = pbir_reader.resolve_page_dir(active, page_id) / "page.json"
    if not ruta.exists():
        return []
    return filter_reader.read_filters(read_json(ruta), scope="pagina")


def _filtros_de_visual(visual: Dict[str, Any]) -> List[Dict[str, Any]]:
    archivo = visual.get("file")
    if not archivo or not Path(archivo).exists():
        return []
    return filter_reader.read_filters(read_json(Path(archivo)), scope="visual")


def _relaciones(session: Session) -> List[Dict[str, Any]]:
    """Relaciones del modelo, para saber que combinaciones existen.

    Se leen del TMDL y no del motor a proposito: `dry_run` tiene que poder
    decir el DAX exacto sin Desktop abierto.
    """
    from horizun_pbi_mcp.pbip import tmdl_reader

    try:
        active = session.require_active_pbip()
    except Exception:                          # noqa: BLE001 - sin pbip activo
        active = None
    if active is not None and active.has_tmdl:
        try:
            return tmdl_reader.read_semantic_model(active).get("relationships") or []
        except Exception as exc:               # noqa: BLE001
            log.warning("No se pudieron leer las relaciones del TMDL: %s", exc)
    try:
        from horizun_pbi_mcp.powerbi import model_reader

        return model_reader.read_model(session).get("relationships") or []
    except Exception:                          # noqa: BLE001 - sin modelo vivo
        return []


def _plan_de_visual_en_pagina(active, pagina: Dict[str, Any],
                              visual: Dict[str, Any],
                              filtros_pagina: List[Dict[str, Any]],
                              relaciones: List[Dict[str, Any]]) -> Dict[str, Any]:
    filtros = filtros_pagina + _filtros_de_visual(visual)
    resumen = filter_reader.resumen(filtros)
    plan = content_query.plan_de_visual(visual, resumen["applied"], relaciones)
    plan["page"] = pagina.get("display_name") or pagina.get("name")
    plan["page_id"] = pagina.get("name")
    plan["filters_applied"] = resumen["applied"]
    plan["filters_untranslated"] = resumen["untranslated"]
    plan["filters_unset"] = resumen["unset"]
    return plan


def resolver_seleccion(session: Session, select: Dict[str, Any]) -> Dict[str, Any]:
    """`select` -> planes de consulta, mas lo que se dejo fuera y por que."""
    if not isinstance(select, dict) or not any(
            select.get(k) for k in ("pages", "visuals", "queries")):
        raise ValidationError(
            "`select` necesita al menos uno de 'pages', 'visuals' o 'queries'.",
            details={"select": select})

    planes: List[Dict[str, Any]] = []
    omitidos: List[Dict[str, Any]] = []
    necesita_pbip = bool(select.get("pages") or select.get("visuals"))
    active = session.require_active_pbip() if necesita_pbip else None
    relaciones = _relaciones(session)

    for referencia in select.get("pages") or []:
        pagina = _pagina_por_nombre(active, referencia)
        filtros_pagina = _filtros_de_pagina(active, pagina["name"])
        for visual in pbir_reader.list_visuals(active, pagina["name"], strict=True):
            plan = _plan_de_visual_en_pagina(active, pagina, visual,
                                             filtros_pagina, relaciones)
            (planes if plan["exportable"] else omitidos).append(plan)

    pedidos = [str(v) for v in (select.get("visuals") or [])]
    if pedidos:
        pendientes = set(pedidos)
        for pagina in pbir_reader.list_pages(active):
            if not pendientes:
                break
            filtros_pagina = _filtros_de_pagina(active, pagina["name"])
            for visual in pbir_reader.list_visuals(active, pagina["name"], strict=True):
                if visual.get("id") not in pendientes:
                    continue
                pendientes.discard(visual["id"])
                plan = _plan_de_visual_en_pagina(active, pagina, visual,
                                                 filtros_pagina, relaciones)
                (planes if plan["exportable"] else omitidos).append(plan)
        for huerfano in sorted(pendientes):
            raise ValidationError(
                f"No existe el visual '{huerfano}' en el informe activo.",
                details={"visual_id": huerfano})

    for spec in select.get("queries") or []:
        planes.append(content_query.plan_declarado(spec, relaciones))

    if not planes:
        raise ValidationError(
            "La seleccion no dejo ninguna consulta que exportar.",
            details={"skipped": [{"title": o.get("title"), "reason": o.get("reason")}
                                 for o in omitidos]})
    if len(planes) > MAX_CONSULTAS:
        raise ValidationError(
            f"La seleccion produce {len(planes)} consultas y el tope es "
            f"{MAX_CONSULTAS}. Acota `pages` o pide visuales concretos.",
            details={"queries": len(planes), "max": MAX_CONSULTAS})
    return {"plans": planes, "skipped": omitidos}


def estado_de_datos(session: Session) -> Dict[str, Any]:
    """Dice si el modelo activo tiene datos cargados, no solo esquema.

    Es la comprobacion que evita el Excel en blanco: un .pbip recien abierto
    sirve el modelo, responde a las consultas y devuelve cero filas.
    """
    from horizun_pbi_mcp.powerbi.adomd_client import AdomdClient

    model = session.require_active_model()
    with AdomdClient(model.connection_string, model.catalog) as client:
        _cols, filas, _t, _e = client.execute_reader(
            "SELECT [TableID], [State], [RefreshedTime] "
            "FROM $SYSTEM.TMSCHEMA_PARTITIONS", max_rows=5_000)
    # `State` 1 es Ready; cualquier otra cosa es una particion sin procesar.
    listas = [f for f in filas if f and str(f[1]) == "1"]
    return {"partitions": len(filas), "ready": len(listas),
            "processed": bool(listas)}


def _modelo_en_vivo(session: Session, *, auto_open: bool,
                    timeout: int = 300) -> List[str]:
    """Deja un modelo en vivo seleccionado. Devuelve avisos."""
    from horizun_pbi_mcp.powerbi.errors import NoActiveModelError

    try:
        session.require_active_model()
        return []
    except NoActiveModelError:
        if not auto_open:
            raise
    active = session.require_active_pbip()
    from horizun_pbi_mcp.powerbi import desktop_discovery, desktop_launcher

    abierto = desktop_launcher.open_pbix(active.pbip_path, timeout=timeout,
                                         reuse_open=True)
    desktop_discovery.select_model(
        session, port=(abierto.instance or {}).get("port"))
    return ["No habia modelo en vivo: se abrio el informe en Power BI Desktop "
            + ("arrancando Desktop." if abierto.launched_by_us
               else "reutilizando la ventana que ya estaba abierta.")]


def _asegurar_modelo(session: Session, *, auto_open: bool,
                     timeout: int = 300) -> List[str]:
    """Modelo en vivo Y con datos. Lo segundo es lo que nadie comprueba."""
    avisos = _modelo_en_vivo(session, auto_open=auto_open, timeout=timeout)
    datos = estado_de_datos(session)
    if not datos["processed"]:
        raise ValidationError(
            "El modelo esta abierto pero SIN DATOS: ninguna de sus "
            f"{datos['partitions']} particiones esta procesada. Un export "
            "ahora saldria en blanco sin un solo error. Refresca el modelo "
            "(pbi_refresh_model) y vuelve a intentarlo.",
            details=datos)
    return avisos


#: `2026-09-11T00:00:00.0000000`, tal como lo devuelve el motor. Es un
#: round-trip de .NET, no algo que un cliente deba leer en un informe.
_FECHA_ISO = re.compile(
    r"^(\d{4}-\d{2}-\d{2})T(\d{2}:\d{2}:\d{2})(?:\.(\d+))?$")


def valor_legible(valor: Any) -> Any:
    """Deja el valor como lo leeria una persona, sin cambiar el dato."""
    if not isinstance(valor, str):
        return valor
    coincidencia = _FECHA_ISO.match(valor.strip())
    if not coincidencia:
        return valor
    fecha, hora, fraccion = coincidencia.groups()
    if hora == "00:00:00" and not (fraccion or "").strip("0"):
        return fecha
    return f"{fecha} {hora}"


def encabezados_legibles(columnas: Sequence[Any]) -> List[str]:
    """`Riesgos[Fase]` -> `Fase`; `[Total]` -> `Total`.

    Si al desnudarlos dos columnas se llamarian igual, TODAS las que chocan
    conservan su forma completa: un encabezado ambiguo en una entrega es peor
    que uno feo.
    """
    desnudos = []
    for columna in columnas:
        texto = str(columna)
        if texto.endswith("]") and "[" in texto:
            desnudos.append(texto[texto.index("[") + 1:-1])
        else:
            desnudos.append(texto)
    repetidos = {n for n in desnudos if desnudos.count(n) > 1}
    return [str(original) if corto in repetidos else corto
            for original, corto in zip(columnas, desnudos)]


def _sin_auxiliares(plan: Dict[str, Any], columnas: List[Any],
                    filas: List[List[Any]]) -> Any:
    """Quita del resultado la medida que solo existia para forzar la relacion.

    El motor devuelve la columna como `[__existe]`. Es andamiaje nuestro: al
    cliente le llega la tabla que pidio, no la que necesitamos para pedirla.
    """
    if not any(m.get("aux") for m in plan.get("measures") or []):
        return columnas, filas
    sobran = {i for i, c in enumerate(columnas)
              if content_query.ALIAS_EXISTENCIA in str(c)}
    if not sobran:
        return columnas, filas
    limpias = [c for i, c in enumerate(columnas) if i not in sobran]
    return limpias, [[v for i, v in enumerate(f) if i not in sobran]
                     for f in filas]


def _ejecutar(session: Session, plan: Dict[str, Any], *,
              max_rows: int) -> Dict[str, Any]:
    resultado = dax_runner.run_dax(session, plan["dax"], max_rows=max_rows)
    columnas, filas = _sin_auxiliares(
        plan, list(resultado.get("columns") or []),
        [list(f) for f in (resultado.get("rows") or [])])
    columnas = encabezados_legibles(columnas)
    filas = [[valor_legible(v) for v in fila] for fila in filas]
    return {"columns": columnas, "rows": filas,
            "row_count": resultado.get("row_count", 0),
            "truncated": bool(resultado.get("truncated")),
            "elapsed_ms": resultado.get("elapsed_ms")}


def _nombre_de_hoja(plan: Dict[str, Any], usados: set) -> str:
    """Excel: 31 caracteres, sin `[]:*?/\\` y sin repetir."""
    base = str(plan.get("title") or plan.get("visual_id") or "Consulta")
    for prohibido in "[]:*?/\\":
        base = base.replace(prohibido, " ")
    base = " ".join(base.split())[:31] or "Consulta"
    candidato, n = base, 2
    while candidato.lower() in usados:
        sufijo = f" {n}"
        candidato, n = base[:31 - len(sufijo)] + sufijo, n + 1
    usados.add(candidato.lower())
    return candidato


def _descripcion_filtros(plan: Dict[str, Any]) -> str:
    partes = []
    for f in plan.get("filters_applied") or plan.get("filters") or []:
        valores = ", ".join(str(v) for v in (f.get("values") or []))
        signo = "no esta en" if f.get("exclude") else "esta en"
        partes.append(f"{f.get('field')} {signo} ({valores})")
    return "; ".join(partes) or "ninguno"


def _descripcion_sin_traducir(plan: Dict[str, Any]) -> str:
    partes = [f"{f.get('field') or 'campo desconocido'}: {f.get('reason')}"
              for f in plan.get("filters_untranslated") or []]
    return "; ".join(partes)


def _hoja_indice(planes: Sequence[Dict[str, Any]],
                 omitidos: Sequence[Dict[str, Any]],
                 avisos: Sequence[str]) -> List[List[Any]]:
    filas: List[List[Any]] = []
    for plan in planes:
        filas.append([
            plan.get("sheet"), plan.get("page") or SIN_DATO,
            plan.get("visual_type"), plan.get("title"),
            plan.get("data", {}).get("row_count"),
            "si" if plan.get("data", {}).get("truncated") else "no",
            _descripcion_filtros(plan),
            _descripcion_sin_traducir(plan) or "ninguno",
            plan.get("dax"),
        ])
    for omitido in omitidos:
        filas.append(["(no exportado)", omitido.get("page") or SIN_DATO,
                      omitido.get("visual_type"), omitido.get("title"),
                      None, SIN_DATO, SIN_DATO, omitido.get("reason"),
                      SIN_DATO])
    for aviso in avisos:
        filas.append(["(aviso)", SIN_DATO, SIN_DATO, aviso, None,
                      SIN_DATO, SIN_DATO, SIN_DATO, SIN_DATO])
    return filas


def _construir_excel(planes: Sequence[Dict[str, Any]],
                     omitidos: Sequence[Dict[str, Any]],
                     avisos: Sequence[str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    indice = wb.active
    indice.title = "Contenido"
    indice.append(["Exportacion", "Horizun PBI MCP"])
    indice.append(["Generado UTC", datetime.now(timezone.utc).isoformat()])
    indice.append([])
    encabezados = ["Hoja", "Pagina", "Tipo de visual", "Titulo", "Filas",
                   "Truncado", "Filtros aplicados", "Filtros NO aplicados",
                   "Consulta DAX"]
    indice.append(encabezados)
    for fila in _hoja_indice(planes, omitidos, avisos):
        indice.append([exporting._json_cell(v) for v in fila])

    cabecera = PatternFill("solid", fgColor="123047")
    acento = PatternFill("solid", fgColor="17A6A6")
    blanco = Font(color="FFFFFF", bold=True)
    for celda in indice[4]:
        celda.fill, celda.font = cabecera, blanco
    indice.freeze_panes = "A5"
    for i, ancho in enumerate([24, 20, 16, 30, 9, 10, 34, 34, 60], 1):
        indice.column_dimensions[get_column_letter(i)].width = ancho

    for plan in planes:
        ws = wb.create_sheet(plan["sheet"])
        datos = plan["data"]
        # El encabezado de contexto va ARRIBA del dato, no en una hoja aparte:
        # quien abre esta hoja tiene que ver con que filtros se saco esto sin
        # tener que ir a buscarlo.
        ws.append([plan.get("title")])
        ws.append(["Pagina", plan.get("page") or SIN_DATO])
        ws.append(["Filtros aplicados", _descripcion_filtros(plan)])
        sin_traducir = _descripcion_sin_traducir(plan)
        if sin_traducir:
            ws.append(["FILTROS NO APLICADOS", sin_traducir])
        if datos["truncated"]:
            ws.append(["AVISO", f"Resultado truncado en {datos['row_count']} filas."])
        ws.append([])
        primera_fila_datos = ws.max_row + 1

        columnas = exporting._unique_headers(datos["columns"]) or ["Sin columnas"]
        ws.append(columnas)
        for fila in datos["rows"]:
            ws.append([exporting._json_cell(v) for v in fila])

        ws["A1"].font = Font(bold=True, size=13, color="123047")
        for celda in ws[primera_fila_datos]:
            celda.fill, celda.font = acento, blanco
            celda.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = f"A{primera_fila_datos + 1}"
        for indice_col, nombre in enumerate(columnas, 1):
            muestra = [str(nombre)] + [
                str(f[indice_col - 1]) for f in datos["rows"][:200]
                if indice_col - 1 < len(f) and f[indice_col - 1] is not None]
            ancho = min(60, max(12, max((len(v) for v in muestra), default=12) + 2))
            ws.column_dimensions[get_column_letter(indice_col)].width = ancho

    wb.properties.title = "Contenido de Power BI"
    wb.properties.creator = "Horizun PBI MCP"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _construir_pdf(planes: Sequence[Dict[str, Any]],
                   omitidos: Sequence[Dict[str, Any]],
                   avisos: Sequence[str], *, titulo: str,
                   max_filas_pdf: int) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    page_size = landscape(A4)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=page_size, rightMargin=14 * mm, leftMargin=14 * mm,
        topMargin=16 * mm, bottomMargin=14 * mm,
        title=exporting._pdf_text(titulo), author="Horizun PBI MCP")
    estilos = getSampleStyleSheet()
    estilos.add(ParagraphStyle(name="HzC", parent=estilos["Title"], fontSize=22,
                               leading=26, alignment=TA_CENTER,
                               textColor=colors.HexColor("#123047")))
    estilos.add(ParagraphStyle(name="HzH", parent=estilos["Heading2"],
                               textColor=colors.HexColor("#123047"),
                               spaceBefore=10, spaceAfter=5))
    estilos.add(ParagraphStyle(name="HzS", parent=estilos["BodyText"],
                               fontSize=8.5, leading=11))
    estilos.add(ParagraphStyle(name="HzHead", parent=estilos["HzS"],
                               textColor=colors.white, fontName="Helvetica-Bold"))

    historia: List[Any] = [
        Paragraph(exporting._pdf_text(titulo), estilos["HzC"]),
        Paragraph(exporting._pdf_text(
            f"Generado {datetime.now(timezone.utc).isoformat()}"), estilos["HzS"]),
        Spacer(1, 6 * mm)]

    def tabla(filas: List[List[Any]], anchos: Optional[List[float]] = None) -> None:
        formateadas = [[Paragraph(exporting._pdf_text(v),
                                  estilos["HzHead"] if i == 0 else estilos["HzS"])
                        for v in fila] for i, fila in enumerate(filas)]
        obj = Table(formateadas, colWidths=anchos, repeatRows=1, hAlign="LEFT")
        obj.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123047")),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B7C5CC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F3F7F8")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        historia.append(obj)

    for aviso in avisos:
        historia.append(Paragraph("- " + exporting._pdf_text(aviso), estilos["HzS"]))

    for plan in planes:
        datos = plan["data"]
        historia.append(Paragraph(exporting._pdf_text(plan.get("title")), estilos["HzH"]))
        contexto = (f"Pagina: {plan.get('page') or '-'} | Filas: "
                    f"{datos['row_count']}{' (truncado)' if datos['truncated'] else ''} "
                    f"| Filtros: {_descripcion_filtros(plan)}")
        historia.append(Paragraph(exporting._pdf_text(contexto), estilos["HzS"]))
        sin_traducir = _descripcion_sin_traducir(plan)
        if sin_traducir:
            historia.append(Paragraph(exporting._pdf_text(
                "FILTROS NO APLICADOS: " + sin_traducir), estilos["HzS"]))
        historia.append(Spacer(1, 2 * mm))

        columnas = list(datos["columns"]) or ["Sin columnas"]
        filas = [columnas] + [list(f) for f in datos["rows"][:max_filas_pdf]]
        ancho_util = page_size[0] - 28 * mm
        tabla(filas, [ancho_util / max(1, len(columnas))] * len(columnas))
        if datos["row_count"] > max_filas_pdf:
            historia.append(Paragraph(exporting._pdf_text(
                f"... {datos['row_count'] - max_filas_pdf} filas mas. El "
                "archivo Excel las lleva todas."), estilos["HzS"]))

    if omitidos:
        historia.append(Paragraph("No exportado", estilos["HzH"]))
        tabla([["Titulo", "Tipo", "Motivo"]] +
              [[o.get("title"), o.get("visual_type"), o.get("reason")]
               for o in omitidos])

    doc.build(historia)
    return buffer.getvalue()


def export_content(session: Session, *, select: Dict[str, Any],
                   format: str = "xlsx", dry_run: bool = False,
                   auto_open: bool = True, max_rows: int = 100_000,
                   max_rows_pdf: int = 40, title: str = "",
                   file_name: str = "") -> Dict[str, Any]:
    """Exporta el contenido que el cliente seleccione a Excel y/o PDF."""
    formato = str(format or "xlsx").strip().lower()
    if formato not in ("xlsx", "pdf", "both"):
        raise ValidationError(
            f"Formato '{format}' no admitido. Usa 'xlsx', 'pdf' o 'both'.",
            details={"format": format})

    resuelto = resolver_seleccion(session, select)
    planes, omitidos = resuelto["plans"], resuelto["skipped"]

    if dry_run:
        return {
            "dry_run": True, "queries": [
                {"title": p.get("title"), "page": p.get("page"),
                 "visual_id": p.get("visual_id"), "dax": p.get("dax"),
                 "filters_applied": p.get("filters_applied") or p.get("filters"),
                 "filters_untranslated": p.get("filters_untranslated") or []}
                for p in planes],
            "skipped": [{"title": o.get("title"), "visual_type": o.get("visual_type"),
                         "reason": o.get("reason")} for o in omitidos],
            "outputs": [], "warnings": [
                "dry_run: no se consulto el motor ni se escribio ningun archivo."],
        }

    avisos = _asegurar_modelo(session, auto_open=auto_open)
    for plan in planes:
        plan["data"] = _ejecutar(session, plan, max_rows=max_rows)

    usados: set = set()
    for plan in planes:
        plan["sheet"] = _nombre_de_hoja(plan, usados)

    titulo = title.strip() or "Contenido del informe de Power BI"
    salidas: List[Dict[str, Any]] = []

    if formato in ("xlsx", "both"):
        payload = _construir_excel(planes, omitidos, avisos)
        destino = exporting._target("content", ".xlsx",
                                    file_name if formato == "xlsx" else "")
        from openpyxl import load_workbook
        reabierto = load_workbook(io.BytesIO(payload), read_only=True)
        esperadas = {"Contenido", *(p["sheet"] for p in planes)}
        if not esperadas.issubset(set(reabierto.sheetnames)):
            reabierto.close()
            raise ValidationError("El Excel generado no conserva todas sus hojas.")
        reabierto.close()
        exporting._publish_new_file(destino, payload)
        salidas.append({"format": "xlsx", "output_path": str(destino),
                        "bytes": len(payload), "verified": True})

    if formato in ("pdf", "both"):
        payload = _construir_pdf(planes, omitidos, avisos, titulo=titulo,
                                 max_filas_pdf=max_rows_pdf)
        destino = exporting._target("content", ".pdf",
                                    file_name if formato == "pdf" else "")
        from pypdf import PdfReader
        paginas = len(PdfReader(io.BytesIO(payload)).pages)
        if paginas < 1:
            raise ValidationError("El PDF generado no tiene paginas.")
        exporting._publish_new_file(destino, payload)
        salidas.append({"format": "pdf", "output_path": str(destino),
                        "bytes": len(payload), "pages": paginas, "verified": True})

    sin_traducir = sum(len(p.get("filters_untranslated") or []) for p in planes)
    if sin_traducir:
        avisos.append(
            f"{sin_traducir} filtro(s) del informe no se pudieron aplicar; "
            "estan declarados en cada hoja. Las cifras pueden no coincidir "
            "con lo que se ve en pantalla.")

    return {
        "dry_run": False, "outputs": salidas,
        "queries": [{"title": p["title"], "sheet": p.get("sheet"),
                     "page": p.get("page"), "rows": p["data"]["row_count"],
                     "truncated": p["data"]["truncated"], "dax": p["dax"],
                     "filters_applied": p.get("filters_applied") or p.get("filters"),
                     "filters_untranslated": p.get("filters_untranslated") or []}
                    for p in planes],
        "skipped": [{"title": o.get("title"), "visual_type": o.get("visual_type"),
                     "reason": o.get("reason")} for o in omitidos],
        "row_total": sum(p["data"]["row_count"] for p in planes),
        "warnings": avisos,
    }
