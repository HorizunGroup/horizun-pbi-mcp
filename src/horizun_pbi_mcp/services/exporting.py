"""Exportaciones verificadas a Excel y PDF.

Los exportadores no escriben dentro del proyecto PBIP. Construyen el artefacto
completo en memoria, lo vuelven a abrir para comprobarlo y solo entonces lo
publican de forma durable bajo ``outputs/``. Los datos proceden exclusivamente
de lectores ya existentes (TOM/TMDL/PBIR/auditoria) y de DAX de solo lectura.
"""
from __future__ import annotations

import io
import html
import hashlib
import json
import math
import os
import shutil
import subprocess
import tempfile
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from horizun_pbi_mcp.config import Session, get_settings
from horizun_pbi_mcp.pbip import pbir_reader, tmdl_reader
from horizun_pbi_mcp.powerbi import dax_runner, model_reader
from horizun_pbi_mcp.powerbi.errors import ValidationError
from horizun_pbi_mcp.services import paths as safe_paths
from horizun_pbi_mcp.services import report_audit
from horizun_pbi_mcp.services.txn import durable_write
from horizun_pbi_mcp.utils.validation import validate_limit


@dataclass
class ExportContext:
    source: str
    model: Dict[str, Any]
    pages: List[Dict[str, Any]] = field(default_factory=list)
    visuals: List[Dict[str, Any]] = field(default_factory=list)
    audit: Optional[Dict[str, Any]] = None
    query_result: Optional[Dict[str, Any]] = None
    warnings: List[str] = field(default_factory=list)


def _normalizar_source(source: str) -> str:
    value = str(source or "auto").strip().lower()
    if value not in ("auto", "live", "pbip"):
        raise ValidationError("source debe ser 'auto', 'live' o 'pbip'.")
    return value


def _collect_context(session: Session, *, source: str, include_report: bool,
                     include_audit: bool, query: str = "",
                     max_rows: int = 100_000) -> ExportContext:
    requested = _normalizar_source(source)
    warnings: List[str] = []

    if requested == "live":
        model = model_reader.read_model(session)
        resolved = "live"
    elif requested == "pbip":
        model = tmdl_reader.read_semantic_model(session.require_active_pbip())
        resolved = "pbip"
    elif session.active_model is not None:
        try:
            model = model_reader.read_model(session)
            resolved = "live"
        except Exception:
            if session.active_pbip is None:
                raise
            model = tmdl_reader.read_semantic_model(session.require_active_pbip())
            resolved = "pbip"
            warnings.append(
                "La sesion live no era utilizable; source='auto' uso el TMDL "
                "del proyecto PBIP activo.")
    elif session.active_pbip is not None:
        model = tmdl_reader.read_semantic_model(session.require_active_pbip())
        resolved = "pbip"
    else:
        raise ValidationError(
            "No hay modelo ni proyecto activo. Selecciona un modelo con "
            "pbi_select_model o abre un PBIP con pbi_open_pbip_project.")

    context = ExportContext(source=resolved, model=model, warnings=warnings)
    active = session.active_pbip
    if include_report:
        if active is None:
            context.warnings.append(
                "No hay proyecto PBIP activo; no se incluyeron paginas ni visuales.")
        else:
            context.pages = pbir_reader.list_pages(active, strict=True)
            for page in context.pages:
                page_visuals = pbir_reader.list_visuals(
                    active, str(page["name"]), strict=True)
                for visual in page_visuals:
                    context.visuals.append({"page": page.get("display_name") or
                                             page.get("name"), **visual})

    if include_audit:
        if active is None:
            context.warnings.append(
                "No hay proyecto PBIP activo; no se incluyo la auditoria integral.")
        else:
            context.audit = report_audit.audit_project(active, model)

    if str(query or "").strip():
        if resolved != "live":
            raise ValidationError(
                "Una consulta DAX requiere source='live' y un modelo Desktop activo.")
        max_rows = int(validate_limit(max_rows, "max_rows", 1_000_000) or 100_000)
        context.query_result = dax_runner.run_dax(
            session, query, max_rows=max_rows, max_bytes=256 * 1024 * 1024,
            export=False)
        if context.query_result.get("truncated"):
            context.warnings.append(
                "La hoja Datos_DAX esta truncada; revisa los limites declarados "
                "en la hoja Resumen antes de sumar o contar.")
    return context


def _target(folder: str, extension: str, file_name: str) -> Path:
    root = (get_settings().outputs_dir / folder).resolve()
    if file_name:
        name = safe_paths.assert_not_path_syntax(
            str(file_name).strip(), kind="nombre de exportacion")
        if len(name) > 160:
            raise ValidationError("El nombre de exportacion supera 160 caracteres.")
        suffix = Path(name).suffix.lower()
        if suffix and suffix != extension:
            raise ValidationError(
                f"El archivo debe terminar en '{extension}', no en '{suffix}'.")
        if not suffix:
            name += extension
    else:
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
        name = f"powerbi_{folder}_{stamp}_{uuid.uuid4().hex[:8]}{extension}"
    target = (root / name).resolve()
    try:
        target.relative_to(root)
    except ValueError as exc:  # defensa adicional a assert_not_path_syntax
        raise ValidationError("La ruta de exportacion sale de outputs/.") from exc
    if target.exists():
        raise ValidationError(
            f"La exportacion ya existe y no se sobrescribe: {target.name}.")
    return target


def _publish_new_file(target: Path, payload: bytes) -> None:
    """Publica de forma durable sin ninguna ventana de sobreescritura.

    La reserva ``O_EXCL`` reclama el nombre exacto. Si la escritura posterior
    falla, solo se elimina el archivo vacio que esta llamada acaba de reservar.
    """
    target.parent.mkdir(parents=True, exist_ok=True)
    try:
        descriptor = os.open(target, os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o600)
    except FileExistsError as exc:
        raise ValidationError(
            f"La exportacion ya existe y no se sobrescribe: {target.name}.") from exc
    os.close(descriptor)
    try:
        durable_write(target, payload)
    except BaseException:
        target.unlink(missing_ok=True)
        raise


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while True:
            chunk = source.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def _json_cell(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (datetime,)):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, default=str,
                           separators=(",", ":"))
    # Excel interpreta estos prefijos como formula incluso si proceden de un
    # nombre o valor remoto. El apostrofo conserva el texto y evita ejecucion.
    text = _neutralizar_formula(str(value))
    # Limite documentado de texto por celda de Excel.
    return text if len(text) <= 32_767 else text[:32_750] + "... [truncado]"


def _record_rows(records: Iterable[Dict[str, Any]], *,
                 preferred: Sequence[str] = (),
                 excluded: Sequence[str] = ()) -> Tuple[List[str], List[List[Any]]]:
    materialized = [r for r in records if isinstance(r, dict)]
    excluded_set = set(excluded)
    keys = set().union(*(r.keys() for r in materialized)) if materialized else set()
    headers = [k for k in preferred if k in keys and k not in excluded_set]
    headers += sorted(k for k in keys if k not in headers and k not in excluded_set)
    return headers, [[_json_cell(record.get(key)) for key in headers]
                     for record in materialized]


def _neutralizar_formula(text: str) -> str:
    """Impide que un texto se convierta en formula al escribirlo en Excel.

    openpyxl marca como formula REAL cualquier cadena que empiece por '='. Los
    otros tres prefijos no crean una formula en el fichero, pero si al pegarlos
    o reintroducirlos en la interfaz, asi que se tratan igual.
    """
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def _unique_headers(headers: Sequence[Any]) -> List[str]:
    """Encabezados validos para Excel Table, estables, unicos y sin formulas.

    Los encabezados de `Datos_DAX` son nombres de columna devueltos por el
    motor: entran igual que un dato y se neutralizan igual que un dato. Sin
    esto, un campo llamado `=...` viajaba crudo y openpyxl lo publicaba como
    formula ejecutable en el libro entregado.
    """
    result: List[str] = []
    counts: Dict[str, int] = {}
    used = set()
    for index, raw in enumerate(headers, 1):
        base = _neutralizar_formula((str(raw).strip() or f"Columna_{index}"))[:240]
        key = base.casefold()
        counts[key] = counts.get(key, 0) + 1
        value = base if counts[key] == 1 else f"{base}_{counts[key]}"
        while value.casefold() in used:
            counts[key] += 1
            value = f"{base}_{counts[key]}"
        result.append(value)
        used.add(value.casefold())
    return result


def _model_sheets(context: ExportContext) -> Dict[str, Tuple[List[str], List[List[Any]]]]:
    model = context.model
    tables = list(model.get("tables") or [])
    columns: List[Dict[str, Any]] = []
    for table in tables:
        for column in table.get("columns") or []:
            columns.append({"table": table.get("name"), **column})
    measures = list(model.get("measures") or [])
    if not measures:
        for table in tables:
            measures.extend({"table": table.get("name"), **measure}
                            for measure in table.get("measures") or [])

    table_rows = []
    for table in tables:
        table_rows.append({k: v for k, v in table.items()
                           if k not in ("columns", "measures")})

    sheets = {
        "Tablas": _record_rows(
            table_rows, preferred=("name", "is_hidden", "column_count",
                                   "measure_count", "description")),
        "Columnas": _record_rows(
            columns, preferred=("table", "name", "data_type", "is_hidden",
                                "description", "source_column")),
        "Medidas": _record_rows(
            measures, preferred=("table", "name", "expression", "dax",
                                 "format_string", "display_folder",
                                 "description")),
        "Relaciones": _record_rows(
            model.get("relationships") or [],
            preferred=("name", "from_table", "from_column", "to_table",
                       "to_column", "from_cardinality", "to_cardinality",
                       "cross_filtering", "is_active")),
    }
    if context.pages:
        sheets["Paginas"] = _record_rows(
            context.pages, preferred=("name", "display_name", "is_active",
                                      "width", "height", "visual_count"))
    if context.visuals:
        sheets["Visuales"] = _record_rows(
            context.visuals, preferred=("page", "id", "type", "title",
                                        "measures", "columns", "fields",
                                        "position"), excluded=("file",))
    if context.audit:
        sheets["Auditoria"] = _record_rows(
            context.audit.get("findings") or context.audit.get("hallazgos") or [],
            preferred=("severity", "domain", "rule", "message", "recommendation",
                       "evidence"))
    if context.query_result:
        headers = [str(c) for c in context.query_result.get("columns") or []]
        sheets["Datos_DAX"] = (
            headers,
            [[_json_cell(v) for v in row]
             for row in context.query_result.get("rows") or []])
    return sheets


def export_excel(session: Session, *, source: str = "auto", query: str = "",
                 include_report: bool = True, include_audit: bool = True,
                 max_rows: int = 100_000, file_name: str = "") -> Dict[str, Any]:
    """Genera y verifica un libro Excel con el modelo, informe y datos DAX."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependencia declarada
        raise ValidationError(
            "Falta openpyxl. Reinstala horizun-pbi-mcp con sus dependencias.") from exc

    context = _collect_context(
        session, source=source, include_report=include_report,
        include_audit=include_audit, query=query, max_rows=max_rows)
    sheets = _model_sheets(context)
    target = _target("excel", ".xlsx", file_name)

    wb = Workbook()
    summary = wb.active
    summary.title = "Resumen"
    audit_score = context.audit.get("score") if context.audit else None
    dax = context.query_result or {}
    summary_rows = [
        ("Exportacion", "Horizun PBI MCP"),
        ("Generado UTC", datetime.now(timezone.utc).isoformat()),
        ("Fuente del modelo", context.source),
        ("Modelo", (context.model.get("model") or {}).get("name")),
        ("Tablas", len(context.model.get("tables") or [])),
        ("Medidas", len(context.model.get("measures") or [])),
        ("Relaciones", len(context.model.get("relationships") or [])),
        ("Paginas", len(context.pages)),
        ("Visuales", len(context.visuals)),
        ("Puntaje auditoria", audit_score),
        ("Filas DAX exportadas", dax.get("row_count")),
        ("DAX truncado", dax.get("truncated")),
    ]
    summary.append(["Indicador", "Valor"])
    for row in summary_rows:
        summary.append([_json_cell(row[0]), _json_cell(row[1])])
    if context.warnings:
        summary.append([])
        summary.append(["Advertencias", "Detalle"])
        for index, warning in enumerate(context.warnings, 1):
            summary.append([index, _json_cell(warning)])

    header_fill = PatternFill("solid", fgColor="123047")
    accent_fill = PatternFill("solid", fgColor="17A6A6")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in [summary]:
        for cell in ws[1]:
            cell.fill, cell.font = header_fill, header_font
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 80

    table_index = 0
    for sheet_name, (headers, rows) in sheets.items():
        if len(headers) > 16_384:
            raise ValidationError(
                f"La hoja '{sheet_name}' supera el limite de 16.384 columnas de Excel.")
        if len(rows) > 1_048_575:
            raise ValidationError(
                f"La hoja '{sheet_name}' supera el limite de filas de Excel.")
        if any(len(row) != len(headers) for row in rows):
            raise ValidationError(
                f"La hoja '{sheet_name}' contiene filas con distinta cantidad "
                "de valores que sus encabezados.")
        ws = wb.create_sheet(sheet_name[:31])
        safe_headers = _unique_headers(headers)
        if not safe_headers:
            safe_headers = ["Sin datos"]
            rows = []
        ws.append(safe_headers)
        for row in rows:
            ws.append(list(row))
        for cell in ws[1]:
            cell.fill, cell.font = accent_fill, header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if rows:
            table_index += 1
            ref = f"A1:{get_column_letter(len(safe_headers))}{len(rows) + 1}"
            table = Table(displayName=f"HzExport{table_index}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)
        for index, header in enumerate(safe_headers, 1):
            sample = [str(header)] + [str(row[index - 1]) for row in rows[:200]
                                      if index - 1 < len(row) and row[index - 1] is not None]
            width = min(60, max(10, max((len(value) for value in sample), default=10) + 2))
            ws.column_dimensions[get_column_letter(index)].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.properties.title = "Exportacion Power BI"
    wb.properties.creator = "Horizun PBI MCP"
    buffer = io.BytesIO()
    wb.save(buffer)
    payload = buffer.getvalue()

    # Verificacion logica ANTES de publicar.
    reopened = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    expected = {"Resumen", *sheets.keys()}
    if not expected.issubset(set(reopened.sheetnames)):
        raise ValidationError("El Excel generado no conserva todas sus hojas.")
    reopened.close()
    _publish_new_file(target, payload)
    # Relectura posterior: comprueba que lo publicado coincide con lo validado.
    if target.read_bytes() != payload:
        target.unlink(missing_ok=True)
        raise ValidationError("El Excel publicado no coincide con el archivo validado.")
    return {
        "output_path": str(target), "format": "xlsx", "source": context.source,
        "sheets": ["Resumen", *sheets.keys()], "bytes": len(payload),
        "verified": True, "warnings": context.warnings,
        "summary": {"tables": len(context.model.get("tables") or []),
                    "measures": len(context.model.get("measures") or []),
                    "pages": len(context.pages), "visuals": len(context.visuals),
                    "dax_rows": dax.get("row_count"),
                    "dax_truncated": dax.get("truncated")},
    }


def _pdf_text(value: Any, limit: int = 500) -> str:
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    text = str(value if value is not None else "")
    if len(text) > limit:
        text = text[:limit - 14] + "... [truncado]"
    # Las fuentes base de PDF usan WinAnsi. Conserva espanol y sustituye solo
    # glifos no representables, evitando cuadrados negros en el render.
    encoded = text.encode("cp1252", errors="replace").decode("cp1252")
    return html.escape(encoded, quote=False)


# Los hallazgos de `report_audit`/`model_audit` identifican SIEMPRE su objeto
# en `object`, nunca en el texto. Un PDF que solo imprime regla y evidencia
# deja siete filas identicas de `measure_possibly_unused` sin decir que medida
# es. Estas dos funciones son las que hacen accionable la tabla de hallazgos.
_MAX_IDS_VISUALES = 4
_ETIQUETA_OBJETO = {
    "model": "Modelo", "report": "Informe", "table": "Tabla",
    "column": "Columna", "measure": "Medida", "relationship": "Relacion",
    "page": "Pagina", "visual": "Visual", "visual_pair": "Visuales",
    "visual_group": "Visuales",
}


def _finding_object_label(objeto: Any,
                          page_names: Optional[Dict[str, str]] = None) -> str:
    """Nombre humano del objeto de un hallazgo: 'Medida Total (Ventas)'."""
    if not isinstance(objeto, dict) or not objeto:
        return ""
    kind = str(objeto.get("kind") or "")
    partes = [_ETIQUETA_OBJETO.get(kind, kind.replace("_", " ")).strip()]

    nombre = objeto.get("name")
    if kind in ("visual", "visual_pair", "visual_group"):
        ids = objeto.get("visuals") or ([objeto["id"]] if objeto.get("id") else [])
        tipo = objeto.get("type")
        if tipo:
            partes.append(str(tipo))
        if ids:
            # Un grupo puede traer los trece visuales de la pagina: la celda
            # crece hasta comerse la fila. El Excel conserva la lista entera.
            visibles = [str(i) for i in ids[:_MAX_IDS_VISUALES]]
            if len(ids) > _MAX_IDS_VISUALES:
                visibles.append(f"(+{len(ids) - _MAX_IDS_VISUALES} mas)")
            partes.append(", ".join(visibles))
    elif kind == "relationship" and not nombre:
        origen, destino = objeto.get("from"), objeto.get("to")
        if origen or destino:
            partes.append(f"{origen or '?'} -> {destino or '?'}")
    elif nombre:
        partes.append(str(nombre))

    if kind == "measure" and objeto.get("table"):
        partes.append(f"({objeto['table']})")

    # `page` es el id interno de la pagina; se traduce al nombre visible.
    pagina = objeto.get("page")
    if pagina and kind != "page":
        visible = (page_names or {}).get(str(pagina)) or str(pagina)
        partes.append(f"- pagina {visible}")
    elif kind == "page":
        visible = nombre or (page_names or {}).get(str(pagina or "")) or pagina
        if visible and not nombre:
            partes.append(str(visible))

    return " ".join(p for p in partes if p).strip()


def _finding_message(finding: Dict[str, Any]) -> str:
    """Texto del hallazgo. Sin `message`, la evidencia se lee como pares.

    Los motores de auditoria no producen `message`: solo `evidence`, que es un
    dict. Volcarlo como JSON deja `{"visual_count": 13}` en un reporte que lee
    una persona. Aqui se aplana a `visual_count: 13`, sin inventar prosa.
    """
    directo = finding.get("message") or finding.get("summary")
    if isinstance(directo, str) and directo.strip():
        return directo
    evidencia = finding.get("evidence")
    if not isinstance(evidencia, dict):
        return "" if evidencia in (None, {}, []) else str(evidencia)
    pares = []
    for clave, valor in evidencia.items():
        if valor is None:
            valor = "(vacio)"
        elif isinstance(valor, bool):        # antes de int: bool es int
            valor = "si" if valor else "no"
        elif isinstance(valor, (list, tuple)):
            valor = ", ".join(str(v) for v in valor)
        elif isinstance(valor, dict):
            valor = json.dumps(valor, ensure_ascii=False, default=str)
        pares.append(f"{clave}: {valor}")
    return "; ".join(pares)


def _capture_files(capture_paths: Optional[Sequence[str]]) -> List[Path]:
    paths: List[Path] = []
    for raw in list(capture_paths or []):
        candidate = Path(str(raw)).expanduser()
        if candidate.is_symlink():
            raise ValidationError(f"La captura no puede ser un enlace: {candidate}.")
        path = candidate.resolve()
        if path.suffix.lower() not in (".png", ".jpg", ".jpeg"):
            raise ValidationError(
                f"Captura no soportada: '{path.name}'. Usa PNG o JPEG.")
        if not path.is_file() or path.is_symlink():
            raise ValidationError(f"La captura no existe o es un enlace: {path}.")
        size = path.stat().st_size
        if size <= 0 or size > 25 * 1024 * 1024:
            raise ValidationError(
                f"La captura '{path.name}' debe pesar entre 1 byte y 25 MB.")
        paths.append(path)
    if len(paths) > 20:
        raise ValidationError("Se admiten como maximo 20 capturas por PDF.")
    return paths


def _png_pixeles_uniformes(data: bytes) -> Optional[bool]:
    """True si todos los pixeles son identicos; None si no se pudo analizar.

    Se decodifica con la biblioteca estandar (PNG 8 bits, sin entrelazar), que
    es lo que emite `pdftoppm -png`. Ante cualquier variante que no se sepa
    leer se devuelve None: no analizar es un resultado honesto; afirmar que la
    pagina tiene contenido sin haberla mirado no lo es.
    """
    import struct
    import zlib

    if not data.startswith(b"\x89PNG\r\n\x1a\n"):
        return None
    ancho = alto = profundidad = tipo = entrelazado = None
    comprimido = bytearray()
    desplazamiento = 8
    while desplazamiento + 8 <= len(data):
        longitud = struct.unpack(">I", data[desplazamiento:desplazamiento + 4])[0]
        clase = data[desplazamiento + 4:desplazamiento + 8]
        cuerpo = data[desplazamiento + 8:desplazamiento + 8 + longitud]
        if clase == b"IHDR" and longitud >= 13:
            (ancho, alto, profundidad, tipo, _compresion, _filtro,
             entrelazado) = struct.unpack(">IIBBBBB", cuerpo[:13])
        elif clase == b"IDAT":
            comprimido += cuerpo
        elif clase == b"IEND":
            break
        desplazamiento += 12 + longitud

    canales = {0: 1, 2: 3, 3: 1, 4: 2, 6: 4}.get(tipo or -1)
    if (ancho is None or canales is None or profundidad != 8
            or entrelazado != 0 or not comprimido):
        return None
    if ancho <= 0 or alto <= 0 or ancho * alto > 40_000_000:
        return None
    try:
        crudo = zlib.decompress(bytes(comprimido))
    except zlib.error:
        return None

    ancho_linea = ancho * canales
    if len(crudo) < alto * (ancho_linea + 1):
        return None

    def paeth(a: int, b: int, c: int) -> int:
        p = a + b - c
        pa, pb, pc = abs(p - a), abs(p - b), abs(p - c)
        return a if pa <= pb and pa <= pc else (b if pb <= pc else c)

    anterior = bytearray(ancho_linea)
    primero: Optional[bytes] = None
    for indice in range(alto):
        inicio = indice * (ancho_linea + 1)
        filtro = crudo[inicio]
        linea = bytearray(crudo[inicio + 1:inicio + 1 + ancho_linea])
        for posicion in range(ancho_linea):
            izquierda = linea[posicion - canales] if posicion >= canales else 0
            arriba = anterior[posicion]
            diagonal = anterior[posicion - canales] if posicion >= canales else 0
            if filtro == 1:
                linea[posicion] = (linea[posicion] + izquierda) & 0xFF
            elif filtro == 2:
                linea[posicion] = (linea[posicion] + arriba) & 0xFF
            elif filtro == 3:
                linea[posicion] = (linea[posicion] + ((izquierda + arriba) >> 1)) & 0xFF
            elif filtro == 4:
                linea[posicion] = (linea[posicion]
                                   + paeth(izquierda, arriba, diagonal)) & 0xFF
            elif filtro != 0:
                return None
        for posicion in range(0, ancho_linea, canales):
            pixel = bytes(linea[posicion:posicion + canales])
            if primero is None:
                primero = pixel
            elif pixel != primero:
                return False
        anterior = linea
    return True


def _render_probe(pdf: bytes, *, capture_pages: Sequence[int] = ()) -> Dict[str, Any]:
    """Renderiza varias paginas del PDF y comprueba que no salen en blanco.

    Mirar siempre la primera pagina no prueba nada del resto: el documento se
    rompe justo donde no se miraba -tablas largas y paginas de capturas-. Se
    renderizan la primera, la ultima y las que llevan captura, con tope.
    """
    from horizun_pbi_mcp import branding
    from pypdf import PdfReader

    configured = branding.env("PDFTOPPM")
    executable = configured or shutil.which("pdftoppm")
    if not executable:
        return {"status": "unavailable", "reason": "pdftoppm no esta instalado"}

    total = len(PdfReader(io.BytesIO(pdf)).pages)
    objetivo = {1, total, *(p for p in capture_pages if 1 <= p <= total)}
    paginas = sorted(objetivo)[:6]

    revisadas: List[Dict[str, Any]] = []
    sin_analizar = 0
    with tempfile.TemporaryDirectory(prefix="horizun_pdf_probe_") as tmp:
        root = Path(tmp)
        source = root / "report.pdf"
        source.write_bytes(pdf)
        for numero in paginas:
            prefix = root / f"page_{numero}"
            try:
                completed = subprocess.run(
                    [executable, "-f", str(numero), "-l", str(numero),
                     "-singlefile", "-png", str(source), str(prefix)],
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30,
                    check=False,
                    creationflags=(0x08000000 if os.name == "nt" else 0))
            except (OSError, subprocess.TimeoutExpired):
                return {"status": "unavailable",
                        "reason": "pdftoppm esta configurado pero no es utilizable"}
            rendered = Path(str(prefix) + ".png")
            if completed.returncode != 0 or not rendered.is_file():
                return {"status": "unavailable",
                        "reason": "pdftoppm esta presente pero no es utilizable",
                        "returncode": completed.returncode}
            data = rendered.read_bytes()
            # Un PNG valido no baja de firma + IHDR + IDAT + IEND. El tamano
            # ya no se usa como proxy de "tiene contenido": una pagina en
            # blanco comprime a muy pocos bytes y de eso se ocupa el analisis
            # de pixeles, que si mira lo que se dibujo.
            if not data.startswith(b"\x89PNG\r\n\x1a\n") or len(data) < 57:
                raise ValidationError(
                    "La prueba visual del PDF produjo una imagen invalida.",
                    details={"page": numero})
            uniforme = _png_pixeles_uniformes(data)
            if uniforme is True:
                raise ValidationError(
                    f"La pagina {numero} del PDF sale en blanco: se renderizo "
                    "sin un solo pixel dibujado.",
                    details={"page": numero, "pages_probed": paginas})
            if uniforme is None:
                sin_analizar += 1
            revisadas.append({"page": numero, "png_bytes": len(data),
                              "blank_check": "no_analizable" if uniforme is None
                              else "con_contenido"})

    return {
        "status": "verified",
        "pages_total": total,
        "pages_rendered": len(revisadas),
        "pages_probed": paginas,
        "first_page_png_bytes": revisadas[0]["png_bytes"] if revisadas else 0,
        "pages_detail": revisadas,
        "blank_check": ("parcial" if sin_analizar else "completo"),
    }


def generate_pdf_report(session: Session, *, report_type: str = "executive",
                        source: str = "auto", title: str = "",
                        capture_paths: Optional[Sequence[str]] = None,
                        include_audit: bool = True, max_findings: int = 50,
                        file_name: str = "") -> Dict[str, Any]:
    """Genera un PDF paginado con resumen, auditoria y capturas verificadas."""
    report_type = str(report_type or "executive").strip().lower()
    if report_type not in ("executive", "technical", "audit"):
        raise ValidationError(
            "report_type debe ser 'executive', 'technical' o 'audit'.")
    max_findings = int(validate_limit(max_findings, "max_findings", 500) or 50)
    captures = _capture_files(capture_paths)
    context = _collect_context(
        session, source=source, include_report=True,
        include_audit=include_audit or report_type == "audit")
    target = _target("pdf", ".pdf", file_name)

    try:
        from pypdf import PdfReader
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import (Image, KeepTogether, PageBreak,
                                        Paragraph, SimpleDocTemplate, Spacer,
                                        Table, TableStyle)
    except ImportError as exc:  # pragma: no cover - dependencias declaradas
        raise ValidationError(
            "Faltan reportlab/pypdf. Reinstala horizun-pbi-mcp.") from exc

    page_size = landscape(A4) if captures else A4
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=page_size, rightMargin=16 * mm, leftMargin=16 * mm,
        topMargin=18 * mm, bottomMargin=16 * mm,
        title=_pdf_text(title or "Reporte Power BI"), author="Horizun PBI MCP")
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="HzTitle", parent=styles["Title"], textColor=colors.HexColor("#123047"),
        alignment=TA_CENTER, fontSize=24, leading=29, spaceAfter=12))
    styles.add(ParagraphStyle(
        name="HzH2", parent=styles["Heading2"], textColor=colors.HexColor("#123047"),
        spaceBefore=12, spaceAfter=7))
    styles.add(ParagraphStyle(
        name="HzSmall", parent=styles["BodyText"], fontSize=8.5, leading=11))
    styles.add(ParagraphStyle(
        name="HzTableHeader", parent=styles["HzSmall"], textColor=colors.white,
        fontName="Helvetica-Bold"))
    story: List[Any] = []
    document_title = _pdf_text(title.strip() if title.strip() else {
        "executive": "Reporte ejecutivo de Power BI",
        "technical": "Documentacion tecnica de Power BI",
        "audit": "Reporte de auditoria de Power BI",
    }[report_type])
    story += [Paragraph(document_title, styles["HzTitle"]),
              Paragraph(_pdf_text(
                  f"Generado {datetime.now(timezone.utc).isoformat()} | "
                  f"Fuente: {context.source}"), styles["HzSmall"]), Spacer(1, 8 * mm)]

    def section(title_text: str) -> None:
        story.append(Paragraph(_pdf_text(title_text), styles["HzH2"]))

    def table(rows: List[List[Any]], widths: Optional[List[float]] = None) -> None:
        formatted = [
            [Paragraph(_pdf_text(value),
                       styles["HzTableHeader"] if row_index == 0 else styles["HzSmall"])
             for value in row]
            for row_index, row in enumerate(rows)
        ]
        obj = Table(formatted, colWidths=widths, repeatRows=1, hAlign="LEFT")
        obj.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123047")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B7C5CC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F3F7F8")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(obj)

    model = context.model
    section("Resumen")
    audit_score = context.audit.get("score") if context.audit else "No incluido"
    table([
        ["Indicador", "Valor"],
        ["Modelo", (model.get("model") or {}).get("name") or "Sin nombre"],
        ["Tablas", len(model.get("tables") or [])],
        ["Medidas", len(model.get("measures") or [])],
        ["Relaciones", len(model.get("relationships") or [])],
        ["Paginas", len(context.pages)], ["Visuales", len(context.visuals)],
        ["Puntaje auditoria", audit_score], ["Capturas", len(captures)],
    ], [55 * mm, 110 * mm])

    if context.warnings:
        section("Advertencias")
        for warning in context.warnings:
            story.append(Paragraph("- " + _pdf_text(warning), styles["HzSmall"]))

    if report_type in ("technical", "executive"):
        section("Modelo semantico")
        model_rows = [["Tabla", "Columnas", "Medidas", "Oculta"]]
        for table_info in model.get("tables") or []:
            model_rows.append([
                table_info.get("name"), table_info.get("column_count", len(table_info.get("columns") or [])),
                table_info.get("measure_count", len(table_info.get("measures") or [])),
                table_info.get("is_hidden", False)])
        table(model_rows, [75 * mm, 28 * mm, 28 * mm, 28 * mm])

    if context.pages:
        section("Paginas del reporte")
        page_rows = [["Pagina", "Activa", "Tamano", "Visuales"]]
        for page in context.pages:
            page_rows.append([
                page.get("display_name") or page.get("name"), page.get("is_active"),
                f"{page.get('width') or '?'} x {page.get('height') or '?'}",
                page.get("visual_count")])
        table(page_rows, [80 * mm, 24 * mm, 38 * mm, 28 * mm])

    findings = []
    if context.audit:
        findings = list(context.audit.get("findings") or
                        context.audit.get("hallazgos") or [])[:max_findings]
        section("Hallazgos de auditoria")
        if findings:
            page_names = {str(p.get("name")): (p.get("display_name") or
                                               p.get("name"))
                          for p in context.pages}
            finding_rows = [["Severidad", "Regla", "Objeto", "Hallazgo",
                             "Recomendacion"]]
            for finding in findings:
                finding_rows.append([
                    finding.get("severity"), finding.get("rule") or finding.get("id"),
                    _finding_object_label(finding.get("object"), page_names),
                    _finding_message(finding), finding.get("recommendation")])
            available = page_size[0] - 32 * mm
            resto = available - 54 * mm
            table(finding_rows, [20 * mm, 34 * mm, resto * 0.30, resto * 0.30,
                                 resto * 0.40])
        else:
            story.append(Paragraph("No se registraron hallazgos.", styles["BodyText"]))

    if report_type == "technical" and model.get("measures"):
        section("Inventario de medidas")
        measure_rows = [["Tabla", "Medida", "Formato", "Carpeta"]]
        for measure in model.get("measures") or []:
            measure_rows.append([
                measure.get("table"), measure.get("name"),
                measure.get("format_string") or measure.get("formatString"),
                measure.get("display_folder") or measure.get("displayFolder")])
        table(measure_rows, [48 * mm, 62 * mm, 32 * mm, 35 * mm])

    if captures:
        story.append(PageBreak())
        section("Capturas de los tableros")
        max_width = page_size[0] - 32 * mm
        max_height = page_size[1] - 55 * mm
        for index, capture in enumerate(captures):
            try:
                reader = ImageReader(str(capture))
                width, height = reader.getSize()
            except Exception as exc:  # noqa: BLE001 - decodificadores de imagen
                raise ValidationError(
                    f"La captura '{capture.name}' no es una imagen valida.") from exc
            if width <= 0 or height <= 0 or width * height > 30_000_000:
                raise ValidationError(
                    f"La captura '{capture.name}' tiene dimensiones no permitidas.")
            scale = min(max_width / width, max_height / height, 1.0)
            image = Image(str(capture), width=width * scale, height=height * scale)
            image.hAlign = "CENTER"
            story.append(KeepTogether([
                Paragraph(_pdf_text(f"Captura {index + 1}: {capture.name}"),
                          styles["HzSmall"]), Spacer(1, 3 * mm), image]))
            if index != len(captures) - 1:
                story.append(PageBreak())

    def footer(canvas, document):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#526872"))
        canvas.drawString(16 * mm, 9 * mm, "Horizun PBI MCP")
        canvas.drawRightString(page_size[0] - 16 * mm, 9 * mm,
                               f"Pagina {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    payload = buffer.getvalue()
    logical = PdfReader(io.BytesIO(payload))
    if len(logical.pages) < 1:
        raise ValidationError("El PDF generado no contiene paginas.")
    # Cada captura ocupa su propia pagina al final del documento (hay un
    # PageBreak antes de la seccion y otro entre capturas), asi que las N
    # ultimas son exactamente las que hay que mirar.
    total_paginas = len(logical.pages)
    paginas_captura = range(max(1, total_paginas - len(captures) + 1),
                            total_paginas + 1) if captures else ()
    render = _render_probe(payload, capture_pages=list(paginas_captura))
    _publish_new_file(target, payload)
    reread = PdfReader(str(target))
    if (len(reread.pages) != len(logical.pages) or
            _sha256_file(target) != hashlib.sha256(payload).hexdigest()):
        target.unlink(missing_ok=True)
        raise ValidationError("El PDF publicado no coincide con el validado.")
    warnings = list(context.warnings)
    if render["status"] == "unavailable":
        warnings.append(
            "PDF verificado logicamente; la verificacion visual con Poppler no "
            "estaba disponible en este runtime.")
    return {
        "output_path": str(target), "format": "pdf", "report_type": report_type,
        "source": context.source, "pages": len(reread.pages), "bytes": len(payload),
        "captures_included": len(captures), "findings_included": len(findings),
        "verified": {"logical": True, "render": render}, "warnings": warnings,
    }
